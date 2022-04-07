# Audit Report v. 0.1
# Timothy Goode (telimektar3)

# Batch reads and processes an auditing tool to determine the average audit score by discipline, and then outputs 
# a file that includes the averages by discipline.
from tkinter import Tk, filedialog
import openpyxl
import os
from datetime import datetime as dt
from openpyxl import Workbook
from collections import Counter as counter

# Creates filename including todays date
mask = '%d%m%Y'
dte = dt.now().strftime(mask)
fname = "/Audit_Averages_Report_{}.xlsx".format(dte)
ffname = "Audit_Averages_Report_{}.xlsx".format(dte)

# Selects a folder containing the audit files necessary to be read using a visual interface
root = Tk() # pointing root to Tk() to use it as Tk() in program.
root.withdraw() # Hides small tkinter window.
root.attributes('-topmost', True) # Opened windows will be active. above all windows despite of selection.
open_folder = filedialog.askdirectory() # Returns opened path as str

    # print(open_folder) # testing that it takes in the selected folder

# Creates list of each Excel file in the selected folder location
files_to_rip = []
for filename in os.listdir(open_folder):
    if filename.endswith(".xlsx") and filename != ffname: # excludes non-.xlsx files and a generated report file
        f = os.path.join(open_folder, filename)
        # checking if it is a file
        if os.path.isfile(f):
            files_to_rip.append(f)
    # print(files_to_rip)

# Imports and structures percentage scores and incomplete scores from the list of files in the folder and stores in new workbook
social_work = []
activity_therapy = []
ics = []
transitional_services = []
forensic_counseling = []
files_processed = 0
social_work_incomplete = []
activity_incomplete = []
ics_incomplete = []
transition_incomplete = []
forensic_incomplete = []



for file in files_to_rip: # looks at each file in the folder
    files_processed += 1
    wb = openpyxl.load_workbook(filename = file, data_only = True) # open the file
    ws = wb['AUDIT TOOL'] # select the necessary sheet
    if ws['E241'].value != "Incomplete":
        social_work.append(ws['E241'].value) # append the necessary values to social work
    else:
        social_work_incomplete.append(file)
    if ws['E242'].value != "Incomplete":
        activity_therapy.append(ws['E242'].value) # append values to activity therapy
    else:
        activity_incomplete.append(file)
    if ws['E243'].value != "Incomplete":
        ics.append(ws['E243'].value) # append values to ICS
    else:
        ics_incomplete.append(file)
    if ws['E244'].value != "Incomplete":
        transitional_services.append(ws['E244'].value) # append values to Transitional Services
    else: 
        transition_incomplete.append(file)
    if ws['E245'].value != "Incomplete":
        forensic_counseling.append(ws['E245'].value)
    else:
        forensic_incomplete.append(file)


item_dict = {"1.1": ws["A19"].value, "1.2": ws["A25"].value, "1.3": ws["A31"].value, "1.4": ws["A37"].value, "1.5": ws["A43"].value, "1.6": ws["A49"].value, "1.7": ws["A55"].value, "2.1": ws["A63"].value, "2.2": ws["A69"].value, "2.3": ws["A75"].value, "2.4": ws["A81"].value, "2.5": ws["A87"].value, "2.6": ws["A93"].value, "2.7": ws["A100"].value, "2.8": ws["A106"].value, "3.1": ws["A114"].value, "3.2": ws["A120"].value, "4.1": ws["A128"].value, "4.2": ws["A134"].value, "4.3": ws["A140"].value, "4.4": ws["A146"].value, "5.1": ws["A154"].value, "5.2": ws["A160"].value, "5.3": ws["A166"].value, "5.4": ws["A172"].value, "5.5": ws["A178"].value, "5.6": ws["A184"].value, "5.7": ws["A190"].value}
   
# Maths the data into averages by discipline
# get sum of all numbers in lists
social_sum = sum(social_work)
activity_sum = sum(activity_therapy)
ics_sum = sum(ics)
transition_sum = sum(transitional_services)
forensic_sum = sum(forensic_counseling)
# get number of numbers in list
social_length = len(social_work)
activity_length = len(activity_therapy)
ics_length = len(ics)
transition_length = len(transitional_services)
forensic_length = len(forensic_counseling)
# get average percentage for each discipline
social_avg = (social_sum / social_length) * 100
activity_avg = (activity_sum / activity_length) * 100
ics_avg = (ics_sum / ics_length) * 100
transition_avg = (transition_sum / transition_length) * 100
forensic_avg = (forensic_sum / forensic_length) * 100

# Pulls "NO" marks from each workbook in selected folder

def get_responsible_clinician(discipline): # this function returns the name of the clinician who matches the passed discipline
    ws = wb['CLINICIANS']
    counselor = ws['B4']
    social_worker = ws['B2']
    activity_therapist = ws['B3']
    ts_counselor = ws['B6']
    fcts = ws['B5']
    discipline_dict = {"Adult Counseling": counselor, "Social Work": social_worker, "Activity Therapy": activity_therapist, "Transitional Services": ts_counselor, "Forensics Clinical Treatment Services": fcts}
    if discipline_dict.get(discipline) != "": 
        ws = wb['AUDIT TOOL']
        return discipline_dict.get(discipline).value
    else:
        ws = wb['AUDIT TOOL']
        return "No clinician listed."

row_number = "" # variable to store the row value of the "NO" response so that it can be used to determine the discipline responsible
audit_form_ranges = [19, 25, 31, 37, 43, 49, 55, 63, 69, 75, 81, 87, 93, 100, 106, 114, 120, 128, 134, 140, 146, 154, 160, 166, 172, 178, 184, 190, 196] # gives the item description cells which also act as bounds for the checks for "NO" "x or X" responses
clinician_item = [] # will store the clinician responsible and the item missed as a list here

def items_missed_getter(cell, cell_value, discipline): # funtion that returns a string of the item missed
    item_missed = ""
    clinician = ""
    inc_count = 0 # this will count for incrementing through the audit form ranges
    if cell_value != None:        # if the cell has a value other than None
        cell_value = cell_value.lower()
        row_number = cell.row     # store the row number without the column identifier
        if cell_value == "x":
            while inc_count != len(audit_form_ranges):
                inc_audit_range = audit_form_ranges[inc_count]
                inc_audit_range_2 = audit_form_ranges[inc_count + 1]
                if row_number > inc_audit_range and row_number < inc_audit_range_2:
                    item_missed = ws['A' + str(audit_form_ranges[inc_count])]
                    clinician = get_responsible_clinician(discipline)
                    inc_count += 1
                    return [item_missed, discipline, clinician]
                else:
                    inc_count += 1


    else:
        return

item_and_clinician = [] # this will store the list of items missed and clinician responsible
for file in files_to_rip: # this loop will create a list that has [item missed, discipline, clinican responsible] sublists
    import openpyxl.cell.cell 
    wb = openpyxl.load_workbook(filename = file, data_only = True) # open the file
    ws = wb['AUDIT TOOL']                                                            # select the necessary sheet
    for row in ws.iter_rows(min_row=20, min_col = 3, max_col=3, max_row=195):        # look at each cell in the "NO" column
        for cell in row:
            row_number = cell.row        # store the row number without the column identifier
            discipline_coordinate = "A" + str(row_number)   # store the coordinate of the discipline responsible
            discipline = ws[discipline_coordinate].value
            cell_value = cell.value
            item_missed = items_missed_getter(cell, cell_value, discipline)
            if item_missed != None:
                if item_missed[0] != "" and item_missed[2] != None:
                    item_missed = [file, item_missed[0], item_missed[1], item_missed[2]]
                    item_and_clinician.append(item_missed)
                else:
                    item_missed = [file, item_missed[0], item_missed[1], "No clinician listed. Please see the following file: " + str(file)] 
                    item_and_clinician.append(item_missed)



# Outputs the data into a new Excel Spreadsheet in the previously selected folder
wb = Workbook()
dest_filename = open_folder + fname
ws = wb.active
ws.title = "Avg Audit Scores"

# assign names to needed cells
ws['A2'] = "Social Work"
ws['B2'] = social_avg
ws['A3'] = "Activity Therapy"
ws['B3'] = activity_avg
ws['A4'] = "Psychology/Counseling"
ws['B4'] = ics_avg
ws['A5'] = "Transitional Services"
ws['B5'] = transition_avg
ws['A6'] = "Forensic Counseling"
ws['B6'] = forensic_avg
ws['C2'] = "Files Processed"
ws['D2'] = files_processed
ws['F2'] = "Social Work files incomplete:"
ws['G2'] = str(social_work_incomplete)
ws['F3'] = "Activity Therapy files incomplete:"
ws['G3'] = str(activity_incomplete)
ws['F4'] = "Psychology/Counseling files incomplete:"
ws['G4'] = str(ics_incomplete)
ws['F5'] = "Transitional Services files incomplete:"
ws['G5'] = str(transition_incomplete)
ws['F6'] = "Forensic Counseling files incomplete:"
ws['G6'] = str(forensic_incomplete)

# Need to output sublists of item_and_clinician to their own two columns
count = 2
for pair in item_and_clinician:
    if pair == None:
        item_and_clinician = item_and_clinician[1:] 
    else:
        ws['I' + str(count)] = pair[0]
        ws['J' + str(count)] = pair[1].value
        ws['K' + str(count)] = pair[2]
        ws['L' + str(count)] = pair[3]
        count += 1

# Create worksheets for each discipline's individual report to pull from
ws_social_work = wb.create_sheet("Social Work")
ws_activity_therapy = wb.create_sheet("Activity Therapy")
ws_ics = wb.create_sheet("Psychology Counseling")
ws_transitional_services = wb.create_sheet("Transitional Services")
ws_forensic_counseling = wb.create_sheet("Forensic Counseling")

report_worksheet_dict = {"Social Work": ws_social_work, "Activity Therapy": ws_activity_therapy, "Adult Counseling": ws_ics, "Transitional Services": ws_transitional_services, "Forensics Clinical Treatment Services": ws_forensic_counseling} # this dict will be used to cycle data to the appropriate sheet

# Process Avg Audit Scores sheet into the individual discipline sheets
data_for_sheet = []
for row in ws.iter_rows(min_row = 2, min_col = 11, max_col = 11): # Iterate through the column reporting what discipline was responsible for a missed item in the current sheet "Avg Audit Scores"
    for cell in row: # For each cell in the column
        if cell.value != None:
            row_number = cell.row # Generate the row number necessary to grab the data in the cells to the left (item missed) and to the right (clinician)
            discipline = cell.value # Get the name of the discipline that missed the item so that it can be put into the correct spreadsheet
            item_missed = ws["J" + str(row_number)].value # Get the value of the item missed
            clinician = ws["L" + str(row_number)].value # Get the name of the clinician who missed the value
            patient_id = ws["I" + str(row_number)].value # Get the file name of the file that the value was missed in
            data_for_sheet.append([discipline, item_missed, clinician, patient_id]) # Append a list of data to be assigned to the appropriate sheet
        else:
            pass


# Messy sorting of the data_for_sheets into their own lists and then put those sorted lists into a new list
social_work_list = []
act_ther_list = []
ics_list = []
tran_list = []
fcts_list = []  
all_discipline_list = [social_work_list, act_ther_list, ics_list, tran_list, fcts_list]

for list in data_for_sheet:
    if list[0] == "Social Work":
        social_work_list.append(list)
    elif list[0] == "Activity Therapy":
        act_ther_list.append(list)
    elif list[0] == "Adult Counseling":
        ics_list.append(list)
    elif list[0] == "Transitional Services":
        tran_list.append(list)
    else:
        fcts_list.append(list)


# Write those sorted lists into their own sheets
for target_discipline in all_discipline_list:
    count = 2
    for list in target_discipline:
        if list == None:
            data_for_sheet = data_for_sheet[1:] 
        else:
            active_spreadsheet = report_worksheet_dict[list[0]] # identify the correct spreadsheet to write to
            active_spreadsheet["A" + str(count)] = list[1] # write the item missed to the an iterated cell in the A column
            active_spreadsheet["B" + str(count)] = list[2] # write the clinician's name to an iterated cell in the B column
            active_spreadsheet["C" + str(count)] = list[3] #write the file name/patient ID to an iterated cell in the C column
            count += 1

def calculate_per_item(discipline, files_processed = files_processed):
    active_spreadsheet = report_worksheet_dict[discipline] # identify the sheet active
    items_missed = []
    for row in active_spreadsheet.iter_rows(min_row = 2, min_col = 1, max_col = 1): # check the item missed
        for cell in row:
            if cell.value != None:
                item_number = cell.value.split(" ", 0)
                item_number = item_number[0]
                items_missed.append(item_number)
    current_amounts = counter(items_missed)
    print(current_amounts)
    sum_items_correct = {}
    for key in current_amounts:
        sum_items_correct[key] = float((files_processed - current_amounts[key])/files_processed) * 100
    return sum_items_correct


for discipline in report_worksheet_dict:
    new_per_item = calculate_per_item(discipline)
    active_spreadsheet = report_worksheet_dict[discipline]
    count = 2
    for item in new_per_item:
        new_item = item.split(" ", 1)
        active_spreadsheet["E" + str(count)] = item + " Percent correct:"
        active_spreadsheet["F" + str(count)] = str(new_per_item[item])
        count += 1

# Saves the workbook
wb.save(open_folder + fname)


# Pop-up stating that the operation has been carried out with "Ok" box
#  NO. Could implement later.